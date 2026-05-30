#!/usr/bin/env python3
"""Generate linear and structured Excel files from Sanskrit morphological analysis."""

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DATA_FILE = Path(__file__).parent / "data" / "sanskrit_verses_2_2_1_to_2_2_9.txt"
OUTPUT_LINEAR = Path(__file__).parent / "sanskrit_verses_linear.xlsx"
OUTPUT_STRUCTURED = Path(__file__).parent / "sanskrit_verses_structured.xlsx"

VERSE_RE = re.compile(r"^verse\s+([\d.]+)\s*$", re.I)
LINE_RE = re.compile(r'^Line\s+(\d+):\s+[""\u201c\u201d](.+?)[""\u201c\u201d]\s*$')
TOKEN_RE = re.compile(r"^(.+?)\s*-\s*$")
ANALYSIS_RE = re.compile(
    r"^(.+?)\s+\((noun|verb|participle|indeclinable(?:\s+\w+)?|pronoun|Preverb|adverb)\s*(?:,\s*(.*?))?\)\s*$"
)
VERB_ROOT_RE = re.compile(r"^√(\S+)\s+\(verb class (\d+)\)\s*$")
PARTICIPLE_RE = re.compile(
    r"^√(\S+)\s+->\s+(\S+)\s+\(participle,\s*(masculine|feminine|neuter)\)\s*$"
)
CANNOT_ANALYSE_RE = re.compile(r"^Cannot analyse\s+(.+)$", re.I)
GRAMMAR_RE = re.compile(r"\[([^\]]+)\]")


def parse_grammar_line(line: str) -> list[str]:
    return [m.group(1).strip() for m in GRAMMAR_RE.finditer(line)]


def parse_analysis_line(line: str) -> dict | None:
    m = PARTICIPLE_RE.match(line)
    if m:
        return {
            "lemma": m.group(2),
            "pos_category": "participle",
            "pos_detail": m.group(3),
            "root": m.group(1),
            "verb_class": None,
            "analysis_text": line,
        }
    m = VERB_ROOT_RE.match(line)
    if m:
        return {
            "lemma": f"√{m.group(1)}",
            "pos_category": "verb",
            "pos_detail": f"class {m.group(2)}",
            "root": m.group(1),
            "verb_class": int(m.group(2)),
            "analysis_text": line,
        }
    m = ANALYSIS_RE.match(line)
    if m:
        lemma, pos_cat, pos_detail = m.group(1), m.group(2), (m.group(3) or "").strip()
        pos_cat = pos_cat.replace("indeclinable adverb", "indeclinable").replace(
            "indeclinable particle", "indeclinable"
        )
        if pos_cat.startswith("indeclinable"):
            detail = pos_cat.replace("indeclinable", "").strip() or pos_detail or "indeclinable"
            return {
                "lemma": lemma,
                "pos_category": "indeclinable",
                "pos_detail": detail or "indeclinable",
                "root": None,
                "verb_class": None,
                "analysis_text": line,
            }
        return {
            "lemma": lemma,
            "pos_category": pos_cat,
            "pos_detail": pos_detail or None,
            "root": None,
            "verb_class": None,
            "analysis_text": line,
        }
    return None


def parse_data(text: str) -> list[dict]:
    verses = []
    current_verse = None
    current_line = None
    current_token = None
    pending_analysis = None

    def flush_token():
        nonlocal current_token, pending_analysis
        if current_token and current_line:
            current_line["tokens"].append(current_token)
        current_token = None
        pending_analysis = None

    def flush_line():
        nonlocal current_line
        flush_token()
        if current_line and current_verse:
            current_verse["lines"].append(current_line)
        current_line = None

    def flush_verse():
        nonlocal current_verse
        flush_line()
        if current_verse:
            verses.append(current_verse)
        current_verse = None

    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            continue

        m = VERSE_RE.match(line)
        if m:
            flush_verse()
            current_verse = {"verse_ref": m.group(1), "lines": []}
            continue

        m = LINE_RE.match(line)
        if m:
            flush_line()
            current_line = {
                "line_num": int(m.group(1)),
                "line_text": m.group(2),
                "tokens": [],
            }
            continue

        m = CANNOT_ANALYSE_RE.match(line)
        if m:
            unresolved_form = m.group(1)
            if current_token is None or current_token["surface"] != unresolved_form:
                flush_token()
                current_token = {
                    "surface": unresolved_form,
                    "marked": False,
                    "analyses": [],
                }
            current_token["analyses"].append(
                {
                    "lemma": None,
                    "pos_category": "unresolved",
                    "pos_detail": None,
                    "root": None,
                    "verb_class": None,
                    "grammar": [],
                    "analysis_text": line,
                    "error": m.group(1),
                }
            )
            pending_analysis = None
            continue

        m = TOKEN_RE.match(line)
        if m and not line.startswith("Line "):
            flush_token()
            surface = m.group(1).strip()
            current_token = {
                "surface": surface.rstrip("*"),
                "marked": surface.endswith("*"),
                "analyses": [],
            }
            pending_analysis = None
            continue

        if line.startswith("[") and pending_analysis is not None:
            pending_analysis["grammar"].extend(parse_grammar_line(line))
            pending_analysis = None
            continue

        analysis = parse_analysis_line(line)
        if analysis and current_token is not None:
            entry = {**analysis, "grammar": []}
            current_token["analyses"].append(entry)
            pending_analysis = entry
            continue

        if line.startswith("[") and current_token is not None and current_token["analyses"]:
            current_token["analyses"][-1]["grammar"].extend(parse_grammar_line(line))
            continue

    flush_verse()
    return verses


def build_linear_rows(verses: list[dict]) -> list[str]:
    rows = []
    for verse in verses:
        rows.append(f"verse {verse['verse_ref']}")
        rows.append("")
        for vline in verse["lines"]:
            rows.append(f'Line {vline["line_num"]}: "{vline["line_text"]}"')
            for token in vline["tokens"]:
                mark = "*" if token["marked"] else ""
                rows.append(f'{token["surface"]}{mark} -')
                for analysis in token["analyses"]:
                    if analysis.get("error"):
                        rows.append(f'Cannot analyse {analysis["error"]}')
                        continue
                    rows.append(analysis["analysis_text"])
                    if analysis["grammar"]:
                        grammar_str = ", ".join(f"[{g}]" for g in analysis["grammar"])
                        rows.append(grammar_str)
            rows.append("")
    return rows


def build_structured_rows(verses: list[dict]) -> list[dict]:
    rows = []
    for verse in verses:
        for vline in verse["lines"]:
            for token in vline["tokens"]:
                if not token["analyses"]:
                    rows.append(
                        {
                            "verse_ref": verse["verse_ref"],
                            "line_num": vline["line_num"],
                            "line_text": vline["line_text"],
                            "token_surface": token["surface"],
                            "token_marked": token["marked"],
                            "analysis_rank": 1,
                            "lemma": None,
                            "pos_category": None,
                            "pos_detail": None,
                            "root": None,
                            "verb_class": None,
                            "grammar_tags": None,
                            "analysis_status": "no_analysis",
                            "full_analysis": None,
                        }
                    )
                    continue
                for rank, analysis in enumerate(token["analyses"], start=1):
                    rows.append(
                        {
                            "verse_ref": verse["verse_ref"],
                            "line_num": vline["line_num"],
                            "line_text": vline["line_text"],
                            "token_surface": token["surface"],
                            "token_marked": token["marked"],
                            "analysis_rank": rank,
                            "lemma": analysis.get("lemma"),
                            "pos_category": analysis.get("pos_category"),
                            "pos_detail": analysis.get("pos_detail"),
                            "root": analysis.get("root"),
                            "verb_class": analysis.get("verb_class"),
                            "grammar_tags": "; ".join(analysis.get("grammar") or []),
                            "analysis_status": "unresolved"
                            if analysis.get("error")
                            else "resolved",
                            "full_analysis": analysis.get("analysis_text"),
                        }
                    )
    return rows


def write_linear_excel(rows: list[str], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Linear Analysis"
    ws.column_dimensions["A"].width = 100

    for i, text in enumerate(rows, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if text.startswith("verse "):
            cell.font = Font(bold=True, size=12)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
        elif text.startswith("Line "):
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="E2EFDA")
        elif text.endswith(" -"):
            cell.font = Font(bold=True, italic=True)
        elif text.startswith("Cannot analyse"):
            cell.font = Font(color="C00000", italic=True)
        elif text.startswith("√") or " (noun," in text or " (verb" in text or " (indeclinable" in text:
            cell.font = Font(italic=True)

    wb.save(path)


def write_structured_excel(rows: list[dict], path: Path) -> None:
    wb = Workbook()

    columns = [
        ("verse_ref", "Verse"),
        ("line_num", "Line #"),
        ("line_text", "Verse Line"),
        ("token_surface", "Token"),
        ("token_marked", "Marked (*)"),
        ("analysis_rank", "Alt #"),
        ("lemma", "Lemma / Root"),
        ("pos_category", "POS Category"),
        ("pos_detail", "POS Detail"),
        ("root", "Dhatu (Root)"),
        ("verb_class", "Verb Class"),
        ("grammar_tags", "Grammar Tags"),
        ("analysis_status", "Status"),
        ("full_analysis", "Raw Analysis"),
    ]

    ws = wb.active
    ws.title = "Morphology"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, (_, header) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(columns, start=1):
            value = row[key]
            if key == "token_marked":
                value = "yes" if value else "no"
            ws.cell(row=row_idx, column=col_idx, value=value)

    widths = {
        "A": 10,
        "B": 8,
        "C": 45,
        "D": 18,
        "E": 10,
        "F": 8,
        "G": 22,
        "H": 16,
        "I": 22,
        "J": 14,
        "K": 10,
        "L": 55,
        "M": 12,
        "N": 40,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    # Training sheet: one row per token with all alternatives as JSON-like text
    ws_train = wb.create_sheet("LLM Training Pairs")
    train_headers = [
        "verse_ref",
        "line_text",
        "token",
        "instruction",
        "input_context",
        "target_analyses",
    ]
    for col_idx, header in enumerate(train_headers, start=1):
        cell = ws_train.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    train_row = 2
    grouped: dict[tuple, list[dict]] = {}
    for row in rows:
        key = (row["verse_ref"], row["line_num"], row["line_text"], row["token_surface"])
        grouped.setdefault(key, []).append(row)

    instruction = (
        "Given a Sanskrit verse line and a token, list all valid morphological analyses "
        "including lemma, part of speech, and grammatical tags."
    )

    for (verse_ref, _line_num, line_text, token_surface), analyses in grouped.items():
        alt_lines = []
        for a in analyses:
            if a["analysis_status"] == "unresolved":
                alt_lines.append(f"UNRESOLVED: {a['full_analysis']}")
            else:
                grammar = a["grammar_tags"] or ""
                alt_lines.append(
                    f"{a['lemma']} ({a['pos_category']}"
                    + (f", {a['pos_detail']}" if a["pos_detail"] else "")
                    + f") | {grammar}"
                )
        ws_train.cell(row=train_row, column=1, value=verse_ref)
        ws_train.cell(row=train_row, column=2, value=line_text)
        ws_train.cell(row=train_row, column=3, value=token_surface)
        ws_train.cell(row=train_row, column=4, value=instruction)
        ws_train.cell(row=train_row, column=5, value=f"Verse line: {line_text}\nToken: {token_surface}")
        ws_train.cell(row=train_row, column=6, value="\n".join(alt_lines))
        train_row += 1

    train_widths = {"A": 10, "B": 45, "C": 18, "D": 50, "E": 50, "F": 70}
    for col, width in train_widths.items():
        ws_train.column_dimensions[col].width = width
    ws_train.freeze_panes = "A2"

    wb.save(path)


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=DATA_FILE)
    parser.add_argument("--output-prefix", type=str, default="sanskrit_verses")
    args = parser.parse_args()

    output_linear = Path(__file__).parent / f"{args.output_prefix}_linear.xlsx"
    structured_dir = Path(__file__).parent / "structured"
    structured_dir.mkdir(exist_ok=True)
    output_structured = structured_dir / f"{args.output_prefix}_structured.xlsx"

    text = args.input.read_text(encoding="utf-8")
    verses = parse_data(text)
    linear_rows = build_linear_rows(verses)
    structured_rows = build_structured_rows(verses)

    write_linear_excel(linear_rows, output_linear)
    write_structured_excel(structured_rows, output_structured)

    print(f"Parsed {len(verses)} verses")
    print(f"Linear rows: {len(linear_rows)}")
    print(f"Structured rows: {len(structured_rows)}")
    print(f"Wrote: {output_linear}")
    print(f"Wrote: {output_structured}")


if __name__ == "__main__":
    main()
